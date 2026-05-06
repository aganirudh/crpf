"""Excel parser — cells emitted as `PageBlock`s, one per non-empty row."""

from __future__ import annotations

import io
import logging

import openpyxl

from pramaan.ingestion.router import PageBlock, hash_text

log = logging.getLogger(__name__)


def parse_xlsx(data: bytes) -> list[PageBlock]:
    blocks: list[PageBlock] = []
    try:
        wb = openpyxl.load_workbook(io.BytesIO(data), data_only=True)
    except Exception as exc:
        log.warning("xlsx parse failed: %s", exc)
        return blocks

    line = 0
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        for row in ws.iter_rows(values_only=True):
            cells = [str(c) for c in row if c is not None and str(c).strip()]
            if not cells:
                continue
            row_text = " | ".join(cells)
            line += 1
            blocks.append(
                PageBlock(
                    page=1,
                    text=row_text,
                    bbox=(0.0, float(line * 15), 600.0, float(line * 15 + 15)),
                    ocr_conf=1.0,
                    extractor="openpyxl",
                    source_text_sha256=hash_text(row_text),
                    extra={"sheet": sheet_name, "cells": cells},
                )
            )
    return blocks
